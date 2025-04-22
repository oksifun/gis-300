import os
import tempfile

from lib.gridfs import get_file_from_gridfs

import settings
from tools.pyoo_helpers import cell_address_to_coordinates, column_letter_to_number, ROW_PARSE_STEP, MAX_COLUMNS


GIS_ERROR_STATUSES = (
    'FMT',
    'INT',
    'SRV',
)


class BaseGISDataImporter:

    USE_OPENPYXL: bool = False

    XLSX_WORKSHEETS: dict = {}

    get_value_openpyxl = lambda sheet, coord: sheet[coord].value
    get_value_pyoo = lambda sheet, coord: sheet[cell_address_to_coordinates(coord)].value

    def _parse_links(self, workbook, get_value=get_value_openpyxl) -> dict:

        parsed_links = {}
        for worksheet_with_links, target_schema in self.XLSX_WORKSHEETS.items():
            if 'links' not in target_schema:
                continue

            for link_target, link_schema in target_schema['links'].items():

                parsed_links[link_target] = []

                linked_schema = self.XLSX_WORKSHEETS[link_schema['from_worksheet']]
                start_row = linked_schema['start_row']
                linked_worksheet = workbook.get_sheet_by_name(link_schema['from_worksheet'])

                linked_match_column = linked_schema['columns'][link_schema['if_match']['linked']]
                linked_value_column = linked_schema['columns'][link_target]
                target_match_column = target_schema['columns'][link_schema['if_match']['target']]

                for row_n in range(start_row, linked_worksheet.max_row):
                    parsed_links[link_target].append({
                        'value': get_value(linked_worksheet, linked_value_column + str(row_n)),
                        'if_column': target_match_column,
                        'match': get_value(linked_worksheet, linked_match_column + str(row_n))
                    })

        return parsed_links

    def import_xlsx_openpyxl(self, import_task):

        from openpyxl import load_workbook, Workbook
        from openpyxl.worksheet.worksheet import Worksheet
        from openpyxl.cell.read_only import EmptyCell

        file = get_file_from_gridfs(file_id=import_task.import_file, raw=True)
        # print('LOADING WORKBOOK', import_task.import_file, 'PLEASE WAIT!')
        book: Workbook = load_workbook(file,
            read_only=True, keep_vba=True,  # VBA-скрипты не выполняются
            data_only=False, keep_links=True)  # сохраняем внешние ссылки

        parsed_links: dict = self._parse_links(book)

        for ws_title, ws_schema in self.XLSX_WORKSHEETS.items():
            worksheet: Worksheet = book.get_sheet_by_name(ws_title)

            for row_num, row in enumerate(worksheet.rows, start=1):
                if row_num < ws_schema['start_row']:  # max_row считает не верно
                    continue

                row_data = {}
                links = {}

                for title, coord in ws_schema['columns'].items():
                    row_data[title] = None
                    for cell in row:  # : ReadOnlyCell, EmptyCell
                        if isinstance(cell, EmptyCell):
                            continue
                        elif cell.coordinate == coord + str(row_num):  # A5
                            row_data[title] = cell.value
                            break  # TODO поиск ячейки без итерации по строке

                for link, conditions in parsed_links.items():
                    for condition in conditions:
                        if worksheet[
                            condition['if_column'] + str(row_num)
                        ].value == condition['match']:
                            links[link] = condition['value']
                            break

                if any(row_data.values()):
                    entry_import_method = \
                        getattr(self, ws_schema['entry_import_method'])
                    entry_import_method(row_data, import_task, links, ws_schema)

    def import_xlsx_pyoo(self, import_task):

        import pyoo  # TODO импортируем если установлено (не импортируем, например, в eve-api)?

        # TODO try-except
        # docker run -d -m 2000M -p 8997:8997 -v /tmp:/tmp -v /home/nizovtsev/tmp/templates:/templates xcgd/libreoffice
        desktop = pyoo.Desktop(settings.LO_DOCKER_HOST, settings.LO_DOCKER_PORT)

        filename, xlsx_content = get_file_from_gridfs(
            file_id=import_task.import_file)

        temp_file = tempfile.NamedTemporaryFile(dir='/tmp/', suffix='.xlsx')
        temp_file.write(xlsx_content)
        temp_file.flush()

        spreadsheet = desktop.open_spreadsheet('/tmp/' + os.path.basename(temp_file.name))

        parsed_links = self._parse_links(spreadsheet, get_value=self.get_value_pyoo)

        for ws_title, ws_schema in self.XLSX_WORKSHEETS.items():
            start_row = ws_schema['start_row']
            sheet = spreadsheet.sheets[ws_title]
            columns = ws_schema['columns']

            fetch_position = 0
            fetch = lambda: list(sheet[fetch_position:fetch_position + ROW_PARSE_STEP, 0:MAX_COLUMNS].values)

            rows_values = fetch()
            while any([value for value in rows_values[-1]]):
                fetch_position += ROW_PARSE_STEP
                rows_values += fetch()

            rows_values = [row for row in rows_values if any([value for value in row])]

            for row_n, row in enumerate(rows_values, start=1):
                if row_n >= start_row:

                    row_data = {}
                    links = {}

                    for column in columns:
                        row_data[column] = row[column_letter_to_number(columns[column])]

                    for link, conditions in parsed_links.items():
                        for condition in conditions:
                            if row[column_letter_to_number(condition['if_column'])] == condition['match']:
                                links[link] = condition['value']
                                break

                    if any(row_data.values()):
                        if ws_schema['entry_import_method'] == 'import_entry_meters_info':
                            getattr(self, ws_schema['entry_import_method'])(row_data, import_task, links, ws_schema, row[42])
                        else:
                            getattr(self, ws_schema['entry_import_method'])(row_data, import_task, links, ws_schema)

        spreadsheet.close()
        temp_file.close()


def is_status_error(status_string: str):
    return True if not status_string or any([status_string.startswith(err) for err in GIS_ERROR_STATUSES]) else False


def get_error_string(status_string: str):
    if not status_string:
        return 'Отсутствует статус обработки'

    words = status_string.split(' ')
    return ' '.join(words[1:]) if is_status_error(words[0]) else status_string


