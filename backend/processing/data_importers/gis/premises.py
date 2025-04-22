from uuid import UUID

from bson import ObjectId

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.gis.models.choices import GisObjectType, GisGUIDStatusType
from app.gis.models.guid import GUID
from app.gis.utils.common import get_time

from processing.data_importers.bulk_writer import BulkWriter


class Storage:

    _instances: dict = {}

    def __new__(cls, fias_guid: str):  # singleton

        if fias_guid not in cls._instances:
            instance = super().__new__(cls)

            print('LOADING FIAS', fias_guid, 'HOUSE DATA...', 'PLEASE WAIT!')
            from app.house.models.house import House
            house: House = House.objects(__raw__={'$or':[
                {'fias_house_guid': fias_guid},
                {'gis_fias': fias_guid},
            ], 'is_deleted': {'$ne': True},
            }).first()
            assert house is not None, \
                f"Дом с идентификатором ФИАС {fias_guid} не найден"

            instance.house_id = house.id
            instance.provider_id = house.get_provider_by_business_type('udo')

            instance.house_guid = GUID.objects(__raw__={
                'tag': GisObjectType.HOUSE,
                'object_id': house.id,  # без provider_id
            }).as_pymongo().first()

            print('LOADING PROVIDER', instance.provider_id,
                'HOUSE', instance.house_id, 'AREAS DATA...', 'PLEASE WAIT!')
            from app.area.models.area import Area
            instance.areas = {area['str_number']: area['_id']
                for area in Area.objects(__raw__={
                    'house._id': house.id,
                    'is_deleted': {'$ne': True},
                }).only('gis_uid', 'str_number').as_pymongo()}

            print('LOADING', len(instance.areas),
                'AREAS GUID DATA...', 'PLEASE WAIT!')
            instance.guids = {guid['object_id']: guid
                for guid in GUID.objects(__raw__={
                    'tag': GisObjectType.AREA,
                    'provider_id': instance.provider_id,
                    'object_id': {'$in': list(instance.areas.values())},
                }).as_pymongo()
            }

            cls._instances[fias_guid] = instance

        return cls._instances[fias_guid]


def load_premises_guids(worksheet: Worksheet, start: int = 0, finish: int = 0):

    if not start:
        start = 1
    if not finish:
        finish = worksheet.max_row

    bulk = BulkWriter(GUID)

    for i, row in enumerate(worksheet.iter_rows(min_row=start, max_row=finish,
            min_col=1, max_col=worksheet.max_column, values_only=True)):
        ind: int = start + i  # с учетом начального сдвига

        # Адрес ОЖФ, HOUSEGUID, FIASGUID, AOGUID (код родительского объекта),
        # ОГРН УО, КПП УО, Краткое наименование УО,
        # Тип дома, Номер дома, Номер корпуса, Номер строения,
        # Признак владения, Признак строения,
        # Номер жилого помещения, Номер жилого блока,
        # Номер нежилого помещения, Номер нежилого блока,
        # Номер комнаты,
        # Уникальный № дома, Уникальный № помещения, Уникальный № комнаты,
        # Кадастровый номер,
        # Общежитие субъекта РФ, Муниципальное общежитие, Тип общежития,
        # Глобальный уникальный идентификатор объекта жилищного фонда,
        # Дата аннулирования объекта, Дата сноса объекта
        address, house_guid, fias_guid, ao_guid, ogrn, kpp, short_name, \
        house_type, house_number, house_block, house_building, \
        owner, building, \
        number, block, non_number, non_block, room_number, \
        house_unique, unique_number, room_unique, cadastral_number, \
        dorm_russian, dorm_municipal, dorm_type, \
        gis_guid, annulled, demolished = row

        storage = Storage(fias_guid)

        gis_guid = UUID(gis_guid)

        if not unique_number and not room_unique:
            assert house_unique, "Уникальный номер дома не определен"
            if storage.house_guid is None:
                storage.house_guid = GUID(
                    # WARN без provider_id и premises_id
                    tag=GisObjectType.HOUSE, object_id=storage.house_id,
                    gis=gis_guid, unique=house_unique,
                    number=house_number, desc=address,
                    status=GisGUIDStatusType.SAVED, saved=get_time()
                ).to_mongo().to_dict()
                bulk.insert(storage.house_guid)
                print(ind, 'CREATING HOUSE', storage.house_id, 'GUID', gis_guid)
            else:
                bulk.update_pk(storage.house_guid['_id'],
                    unset=['provider_id', 'record_id', 'transport', 'error'],
                    gis=gis_guid, unique=house_unique,
                    status=GisGUIDStatusType.SAVED, saved=get_time())
                print(ind, 'UPDATING HOUSE', storage.house_id, 'GUID', gis_guid)
            continue

        is_living: bool = number is not None and number.strip != ''  # 0?
        str_number: str = number if is_living else non_number

        if annulled is not None:
            print(ind, 'SKIPPING ANNULLED AREA NUMBER', str_number)
            continue

        area_id: ObjectId = storage.areas.get(str_number)
        if area_id is None:
            print(ind, 'AREA UNIQUE', unique_number,
                'NUMBER', str_number, 'NOT FOUND!')
            continue

        guid: dict = storage.guids.get(area_id)  # WARN после HouseAreas.of
        if guid is None:
            assert storage.provider_id is not None, \
                "Идентификатор управляющей организации не определен"
            assert storage.house_id is not None, \
                "Идентификатор дома не определен"
            guid: dict = GUID(
                tag=GisObjectType.AREA, object_id=area_id,
                provider_id=storage.provider_id,
                premises_id=storage.house_id,
                gis=gis_guid, unique=unique_number, number=str_number,
                desc=f"{'кв.' if is_living else 'пом.'} {str_number}",
            ).to_mongo().to_dict()
            bulk.insert(guid)
            print(ind, 'CREATING AREA', area_id, 'GUID', gis_guid)
        else:
            assert guid.get('premises_id') is None or \
                    storage.house_id == guid['premises_id'], \
                "Неверный идентификатор дома в данных ГИС ЖКХ помещения"
            bulk.update_pk(guid['_id'],
                unset=['record_id', 'transport', 'error'],  # WARN удаляем поля
                premises_id=storage.house_id,
                gis=gis_guid, unique=unique_number,
                status=GisGUIDStatusType.SAVED, saved=get_time())
            print(ind, 'UPDATING AREA', area_id, 'GUID', gis_guid)

    bulk.write()


if __name__ == '__main__':

    from mongoengine_connections import register_mongoengine_connections
    register_mongoengine_connections()

    f = 'processing/data_importers/gis/data/Объекты_жилищного_фонда_от_13_04_2023_12_13.xlsx'

    wb: Workbook = load_workbook(
        f,
        read_only=True, keep_vba=False,  # VBA-скрипты не выполняются
        data_only=True, keep_links=False  # сохраняем внешние ссылки
    )

    ws: Worksheet = wb.active

    load_premises_guids(
        ws,
        start=3,
        finish=7983  # TODO ws.max_row - дает неверный результат
    )
